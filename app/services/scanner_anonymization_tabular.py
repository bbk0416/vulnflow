"""CSV and XLSX anonymization while preserving scanner column contracts."""
from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook, load_workbook

from app.services.finding_import_common import _clean_cell, _header_key
from app.services.finding_import_tabular import _decode_csv, _xlsx_zip_preflight
from app.services.finding_imports import auto_map_headers
from app.services.scanner_anonymization_common import AliasVault, classify_identity, safe_scalar

_IDENTITY_FIELDS = {"asset_name", "ip_address", "fqdn", "asset_id", "owner", "environment"}
_PRODUCT_FIELDS = {"product", "component", "product_version", "component_version"}
_FREE_TEXT_FIELDS = {"notes"}
_IDENTITY_HEADER_WORDS = ("host", "asset", "target", "ip", "fqdn", "dns", "uuid", "guid", "mac", "owner", "user", "email", "computer", "domain", "customer", "company", "organization", "tenant", "project", "site", "location")
_FREE_TEXT_HEADER_WORDS = ("description", "summary", "synopsis", "solution", "output", "result", "detail", "note", "technical")
_PRODUCT_HEADER_WORDS = ("product", "plugin", "nvt", "service", "component", "version", "cpe")


def _field_for_header(header: str, mapping: dict[str, str]) -> str:
    for field, source in mapping.items():
        if source == header:
            return field
    key = _header_key(header)
    if any(word in key for word in _IDENTITY_HEADER_WORDS):
        return "asset_name"
    if any(word in key for word in _FREE_TEXT_HEADER_WORDS):
        return "notes"
    if any(word in key for word in _PRODUCT_HEADER_WORDS):
        return "product"
    return ""


def _sanitize_value(vault: AliasVault, field: str, value: Any, *, profile: str) -> str:
    text = _clean_cell(value)
    if not text:
        return ""
    if field in _FREE_TEXT_FIELDS:
        vault.source_tokens.add(text) if len(text) >= 4 else None
        return "Redacted by VulnFlow anonymizer."
    if field == "ip_address":
        return vault.alias("ip", text)
    if field == "fqdn":
        return vault.alias("fqdn", text)
    if field == "asset_id":
        return vault.alias("uuid" if "-" in text else "asset", text)
    if field == "owner":
        return vault.alias("email" if "@" in text else "owner", text)
    if field == "environment":
        return vault.alias("environment", text)
    if field == "asset_name":
        return vault.alias(classify_identity(text), text)
    if field in _PRODUCT_FIELDS:
        if profile == "strict":
            category = "version" if "version" in field else "component" if field == "component" else "product"
            return vault.alias(category, text)
        return text
    if safe_scalar(text):
        return text
    if not field or profile == "strict":
        vault.source_tokens.add(text) if len(text) >= 4 else None
        return "Redacted by VulnFlow anonymizer."
    return text


def sanitize_csv(content: bytes, *, vault: AliasVault, profile: str) -> tuple[bytes, dict[str, Any]]:
    text, encoding = _decode_csv(content)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        raise ValueError("파일에 헤더가 없습니다.")
    headers = [str(value) for value in rows[0]]
    mapping = auto_map_headers(headers)
    output = io.StringIO(newline="")
    writer = csv.writer(output, dialect)
    writer.writerow(headers)
    for values in rows[1:]:
        padded = list(values[: len(headers)]) + [""] * max(0, len(headers) - len(values))
        writer.writerow([
            _sanitize_value(vault, _field_for_header(header, mapping), padded[index], profile=profile)
            for index, header in enumerate(headers)
        ])
    payload = ("\ufeff" + output.getvalue()).encode("utf-8")
    return payload, {"source_encoding": encoding, "delimiter": getattr(dialect, "delimiter", ","), "rows": max(0, len(rows) - 1)}


def sanitize_xlsx(content: bytes, *, vault: AliasVault, profile: str) -> tuple[bytes, dict[str, Any]]:
    _xlsx_zip_preflight(content)
    source = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
    target = Workbook(write_only=False)
    target.remove(target.active)
    sheet_count = 0
    row_count = 0
    try:
        for source_sheet in source.worksheets:
            rows = list(source_sheet.iter_rows(values_only=True))
            nonempty = next((index for index, row in enumerate(rows) if any(_clean_cell(value) for value in row)), None)
            if nonempty is None:
                continue
            sheet_count += 1
            target_sheet = target.create_sheet(title=f"Sheet{sheet_count}")
            headers = [_clean_cell(value) or f"column_{index + 1}" for index, value in enumerate(rows[nonempty])]
            mapping = auto_map_headers(headers)
            target_sheet.append(headers)
            for values in rows[nonempty + 1:]:
                if not any(_clean_cell(value) for value in values):
                    continue
                padded = list(values[: len(headers)]) + [None] * max(0, len(headers) - len(values))
                target_sheet.append([
                    _sanitize_value(vault, _field_for_header(header, mapping), padded[index], profile=profile)
                    for index, header in enumerate(headers)
                ])
                row_count += 1
        if not target.worksheets:
            raise ValueError("XLSX에 데이터가 있는 시트가 없습니다.")
        target.properties.creator = "VulnFlow anonymizer"
        target.properties.lastModifiedBy = "VulnFlow anonymizer"
        target.calculation.fullCalcOnLoad = False
        output = io.BytesIO()
        target.save(output)
        return output.getvalue(), {"sheets": sheet_count, "rows": row_count, "values_only": True}
    finally:
        source.close()
        target.close()


__all__ = ["sanitize_csv", "sanitize_xlsx"]
