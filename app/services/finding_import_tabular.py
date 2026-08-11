"""Tabular CSV and XLSX parsing for finding imports."""
from __future__ import annotations

import csv
import io
import zipfile
from typing import Any

from openpyxl import load_workbook

from app.core.settings import MAX_CSV_ROWS
from app.services.finding_import_common import _clean_cell, _unique_headers

def _decode_csv(content: bytes) -> tuple[str, str]:
    if b"\x00" in content:
        raise ValueError("CSV 파일에 허용되지 않는 NUL 문자가 있습니다.")
    errors: list[str] = []
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
    raise ValueError("CSV 문자 인코딩을 확인할 수 없습니다. UTF-8 또는 CP949로 저장하세요.")

def _csv_rows(content: bytes) -> dict[str, Any]:
    text, encoding = _decode_csv(content)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect, strict=True)
    try:
        raw_headers = next(reader)
    except StopIteration as exc:
        raise ValueError("파일에 헤더가 없습니다.") from exc
    except csv.Error as exc:
        raise ValueError(f"CSV 형식 오류: {exc}") from exc
    headers = _unique_headers(raw_headers)
    rows: list[dict[str, str]] = []
    source_rows: list[int] = []
    try:
        while True:
            source_row = reader.line_num + 1
            try:
                values = next(reader)
            except StopIteration:
                break
            if not any(_clean_cell(value) for value in values):
                continue
            overflow = values[len(headers):]
            if any(_clean_cell(value) for value in overflow):
                raise ValueError(
                    f"CSV 행 {source_row}의 열 수가 헤더보다 많습니다. "
                    "구분자 또는 따옴표를 확인하세요."
                )
            if len(rows) >= MAX_CSV_ROWS:
                raise ValueError(f"가져오기는 최대 {MAX_CSV_ROWS:,}행까지 지원합니다.")
            padded = list(values[: len(headers)]) + [""] * max(0, len(headers) - len(values))
            rows.append({header: _clean_cell(padded[index]) for index, header in enumerate(headers)})
            source_rows.append(source_row)
    except csv.Error as exc:
        raise ValueError(f"CSV 형식 오류: {exc}") from exc
    return {
        "headers": headers,
        "rows": rows,
        "source_rows": source_rows,
        "metadata": {"encoding": encoding, "delimiter": getattr(dialect, "delimiter", ",")},
    }

def _xlsx_zip_preflight(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > 2000:
                raise ValueError("XLSX 내부 파일 수가 비정상적으로 많습니다.")
            total_size = sum(info.file_size for info in infos)
            if total_size > 80 * 1024 * 1024:
                raise ValueError("XLSX 압축 해제 크기는 최대 80MB입니다.")
            for info in infos:
                if info.compress_size and info.file_size / info.compress_size > 1000:
                    raise ValueError("XLSX 압축률이 비정상적으로 높습니다.")
    except zipfile.BadZipFile as exc:
        raise ValueError("올바른 XLSX 파일이 아닙니다.") from exc

def _xlsx_rows(content: bytes) -> dict[str, Any]:
    _xlsx_zip_preflight(content)
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
    except Exception as exc:  # openpyxl exposes several parser-specific exceptions
        raise ValueError(f"XLSX 파일을 읽을 수 없습니다: {exc}") from exc
    try:
        selected = None
        header_values: list[Any] = []
        data_iterator = None
        header_row_number = 0
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            for physical_row, row in enumerate(iterator, start=1):
                if any(_clean_cell(value) for value in row):
                    selected = worksheet
                    header_values = list(row)
                    header_row_number = physical_row
                    last_header_index = max(
                        index for index, value in enumerate(header_values, start=1) if _clean_cell(value)
                    )
                    header_values = header_values[:last_header_index]
                    data_iterator = iterator
                    break
            if selected is not None:
                break
        if selected is None or data_iterator is None:
            raise ValueError("XLSX에 데이터가 있는 시트가 없습니다.")
        if len(header_values) > 200:
            raise ValueError("XLSX는 최대 200개 열까지 지원합니다.")
        headers = _unique_headers(header_values)
        rows: list[dict[str, str]] = []
        source_rows: list[int] = []
        source_row = header_row_number
        for values in data_iterator:
            source_row += 1
            if not any(_clean_cell(value) for value in values):
                continue
            overflow = list(values[len(headers):])
            if any(_clean_cell(value) for value in overflow):
                raise ValueError(
                    f"XLSX 행 {source_row}의 열 수가 헤더보다 많습니다. "
                    "헤더 범위 또는 시트 구조를 확인하세요."
                )
            if len(rows) >= MAX_CSV_ROWS:
                raise ValueError(f"가져오기는 최대 {MAX_CSV_ROWS:,}행까지 지원합니다.")
            padded = list(values[: len(headers)]) + [None] * max(0, len(headers) - len(values))
            rows.append({header: _clean_cell(padded[index]) for index, header in enumerate(headers)})
            source_rows.append(source_row)
        return {
            "headers": headers,
            "rows": rows,
            "source_rows": source_rows,
            "metadata": {"sheet_name": selected.title, "sheet_count": len(workbook.sheetnames)},
        }
    finally:
        workbook.close()

