from __future__ import annotations

import io
import re
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
import pytest

import app.main as main
from app.core.storage import apply_import_batch, get_source_reconciliation, init_db, list_assets, list_findings
from app.services.asset_identity import extract_asset_identifiers, normalize_asset_identifier
from app.services.finding_imports import (
    auto_map_headers,
    create_preview_session,
    load_preview_session,
    map_import_rows,
    parse_import_file,
)


def _csrf(client: TestClient) -> str:
    response = client.get("/upload")
    assert response.status_code == 200
    return client.cookies.get(main.CSRF_COOKIE)


def test_cp949_csv_and_korean_header_auto_mapping():
    content = "제품명,CVE,호스트명,IP 주소,CVSS\n웹서버,CVE-2026-12345,server-a,10.0.0.1,9.8\n".encode("cp949")
    parsed = parse_import_file(content, filename="진단결과.csv")
    assert parsed["detected_format"] == "csv"
    assert parsed["metadata"]["encoding"] == "cp949"
    assert parsed["mapping"]["product"] == "제품명"
    assert parsed["mapping"]["cve_id"] == "CVE"
    assert parsed["mapping"]["asset_name"] == "호스트명"

    with pytest.raises(ValueError, match="열 수가 헤더보다 많습니다"):
        parse_import_file(
            b"product,cve_id,asset_name,cvss\nOpenSSL,CVE-2026-12345,host-a,9.8,EXTRA\n",
            filename="ragged.csv",
        )
    with pytest.raises(ValueError, match="CSV 형식 오류"):
        parse_import_file(
            b'product,cve_id,asset_name,notes\nOpenSSL,CVE-2026-12345,host-a,"unterminated\n',
            filename="broken-quotes.csv",
        )


def test_xlsx_first_nonempty_sheet_and_cve_expansion():
    workbook = Workbook()
    empty = workbook.active
    empty.title = "안내"
    sheet = workbook.create_sheet("진단결과")
    sheet.append(["제품", "CVE", "자산명", "CVSS"])
    sheet.append(["OpenSSL", "CVE-2026-11111, CVE-2026-22222", "api-1", 9.1])
    buffer = io.BytesIO()
    workbook.save(buffer)
    parsed = parse_import_file(buffer.getvalue(), filename="result.xlsx")
    assert parsed["detected_format"] == "xlsx"
    assert parsed["metadata"]["sheet_name"] == "진단결과"
    mapped, source_rows, errors = map_import_rows(parsed["rows"], parsed["source_rows"], parsed["mapping"])
    assert errors == []
    assert [row["cve_id"] for row in mapped] == ["CVE-2026-11111", "CVE-2026-22222"]
    assert source_rows == [2, 2]

    overflow = Workbook()
    overflow_sheet = overflow.active
    overflow_sheet.append(["product", "cve_id"])
    overflow_sheet.append(["OpenSSL", "CVE-2026-12345", "EXTRA"])
    overflow_buffer = io.BytesIO()
    overflow.save(overflow_buffer)
    with pytest.raises(ValueError, match="XLSX 행 2의 열 수가 헤더보다 많습니다"):
        parse_import_file(overflow_buffer.getvalue(), filename="ragged.xlsx")



def test_csv_multiline_records_preserve_physical_source_rows():
    content = (
        'product,cve_id,notes\n'
        'One,CVE-2026-77704,"line one\nline two"\n'
        'Two,CVE-2026-77705,plain\n'
    ).encode()
    parsed = parse_import_file(content, filename="multiline.csv", format_hint="csv")
    assert parsed["source_rows"] == [2, 4]


def test_xlsx_leading_blank_rows_preserve_physical_source_rows_and_overflow_location():
    workbook = Workbook()
    sheet = workbook.active
    for _ in range(4):
        sheet.append([])
    sheet.append(["product", "cve_id"])
    sheet.append(["One", "CVE-2026-77706"])
    sheet.append(["Two", "CVE-2026-77707"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    parsed = parse_import_file(buffer.getvalue(), filename="blank-prefix.xlsx", format_hint="xlsx")
    assert parsed["source_rows"] == [6, 7]

    overflow = Workbook()
    overflow_sheet = overflow.active
    for _ in range(4):
        overflow_sheet.append([])
    overflow_sheet.append(["product", "cve_id"])
    overflow_sheet.append(["One", "CVE-2026-77708"])
    overflow_sheet.append(["Two", "CVE-2026-77709", "EXTRA"])
    overflow_buffer = io.BytesIO()
    overflow.save(overflow_buffer)
    with pytest.raises(ValueError, match="XLSX 행 7의 열 수가 헤더보다 많습니다"):
        parse_import_file(overflow_buffer.getvalue(), filename="blank-prefix-overflow.xlsx", format_hint="xlsx")

def test_nessus_adapter_extracts_cves_and_reports_non_cve_plugins():
    payload = b"""<?xml version='1.0'?>
<NessusClientData_v2><Report name='demo'><ReportHost name='10.0.0.8'>
<HostProperties><tag name='host-ip'>10.0.0.8</tag><tag name='host-fqdn'>web.example.test</tag></HostProperties>
<ReportItem port='443' svc_name='https' pluginID='1001' pluginName='OpenSSL issue'>
<cve>CVE-2026-30001</cve><cve>CVE-2026-30002</cve><cvss3_base_score>9.8</cvss3_base_score>
<synopsis>Critical TLS issue</synopsis><solution>Upgrade OpenSSL</solution>
</ReportItem>
<ReportItem port='0' svc_name='general' pluginID='1002' pluginName='Informational item'><synopsis>No CVE</synopsis></ReportItem>
</ReportHost></Report></NessusClientData_v2>"""
    parsed = parse_import_file(payload, filename="scan.nessus")
    assert parsed["detected_format"] == "nessus"
    assert len(parsed["rows"]) == 2
    assert len(parsed["source_errors"]) == 1
    assert parsed["rows"][0]["asset_name"] == "web.example.test"
    assert parsed["rows"][0]["patch_available"] == "1"
    mapped, source_rows, errors = map_import_rows(parsed["rows"], parsed["source_rows"], parsed["mapping"])
    assert len(mapped) == 2 and errors == [] and source_rows == [1, 1]

    hostname_only = b"""<?xml version='1.0'?>
<NessusClientData_v2><Report name='demo'><ReportHost name='db01'>
<HostProperties></HostProperties>
<ReportItem port='0' svc_name='general' pluginID='2001' pluginName='Hostname only'>
<cve>CVE-2026-30003</cve></ReportItem>
</ReportHost></Report></NessusClientData_v2>"""
    hostname_parsed = parse_import_file(hostname_only, filename="hostname-only.nessus")
    assert hostname_parsed["rows"][0]["asset_name"] == "db01"
    assert hostname_parsed["rows"][0]["ip_address"] == ""
    identifiers = extract_asset_identifiers(hostname_parsed["rows"][0], scanner_source="nessus")
    assert any(item["identifier_type"] == "HOSTNAME" and item["normalized_value"] == "db01" for item in identifiers)

    malformed_host_ip = b"""<?xml version='1.0'?>
<NessusClientData_v2><Report name='demo'><ReportHost name='app-host'>
<HostProperties><tag name='host-ip'>db01</tag></HostProperties>
<ReportItem port='0' svc_name='general' pluginID='2002' pluginName='Bad host-ip'>
<cve>CVE-2026-30004</cve></ReportItem>
</ReportHost></Report></NessusClientData_v2>"""
    malformed_parsed = parse_import_file(malformed_host_ip, filename="malformed-host-ip.nessus")
    assert malformed_parsed["rows"][0]["ip_address"] == ""
    assert malformed_parsed["rows"][0]["asset_name"] == "app-host"
    assert any("host-ip" in warning for warning in malformed_parsed["parser_warnings"])
    extract_asset_identifiers(malformed_parsed["rows"][0], scanner_source="nessus")


def test_openvas_csv_is_detected_and_mapped():
    content = (
        "IP,Hostname,Port,NVT Name,CVEs,CVSS,Summary\n"
        "10.0.0.9,db-1,5432/tcp,PostgreSQL issue,CVE-2026-44444,8.8,Upgrade required\n"
    ).encode()
    parsed = parse_import_file(content, filename="greenbone.csv")
    assert parsed["detected_format"] == "openvas_csv"
    assert parsed["scanner_source_suggestion"] == "openvas"
    assert parsed["mapping"]["product"] == "product"
    assert parsed["mapping"]["cve_id"] == "cve_id"
    assert parsed["mapping"]["ip_address"] == "ip_address"

    hostname_content = (
        "Host,NVT Name,CVEs,CVSS\n"
        "db01,Hostname-only target,CVE-2026-44445,7.1\n"
        "dead.beef,Hex-looking FQDN,CVE-2026-44446,7.2\n"
    ).encode()
    hostname_parsed = parse_import_file(hostname_content, filename="greenbone-hostnames.csv", format_hint="openvas")
    assert [row["ip_address"] for row in hostname_parsed["rows"]] == ["", ""]
    assert hostname_parsed["rows"][0]["asset_name"] == "db01"
    assert hostname_parsed["rows"][1]["asset_name"] == "dead.beef"
    first_ids = extract_asset_identifiers(hostname_parsed["rows"][0], scanner_source="openvas")
    second_ids = extract_asset_identifiers(hostname_parsed["rows"][1], scanner_source="openvas")
    assert any(item["identifier_type"] == "HOSTNAME" and item["normalized_value"] == "db01" for item in first_ids)
    assert any(item["identifier_type"] == "FQDN" and item["normalized_value"] == "dead.beef" for item in second_ids)



def test_openvas_xml_adapter_extracts_result():
    payload = b"""<?xml version='1.0'?><get_reports_response><report><report><results>
    <result id='r1'><name>TLS weakness</name><host>10.0.0.30<hostname>tls.example.test</hostname></host>
    <port>443/tcp</port><nvt oid='1.3.6'><name>TLS weakness</name><cve>CVE-2026-65001</cve><cvss_base>7.5</cvss_base></nvt>
    <description>Upgrade the affected TLS library.</description></result>
    </results></report></report></get_reports_response>"""
    parsed = parse_import_file(payload, filename="report.xml")
    assert parsed["detected_format"] == "openvas_xml"
    assert parsed["rows"][0]["cve_id"] == "CVE-2026-65001"
    assert parsed["rows"][0]["asset_name"] == "tls.example.test"
    assert parsed["rows"][0]["ip_address"] == "10.0.0.30"

    hostname_payload = b"""<?xml version='1.0'?><get_reports_response><report><report><results>
    <result id='r2'><name>Hostname target</name><host>db01</host>
    <port>443/tcp</port><nvt oid='1.3.7'><name>Hostname target</name><cve>CVE-2026-65002</cve><cvss_base>7.1</cvss_base></nvt>
    </result></results></report></report></get_reports_response>"""
    hostname_parsed = parse_import_file(hostname_payload, filename="hostname-report.xml")
    assert hostname_parsed["rows"][0]["asset_name"] == "db01"
    assert hostname_parsed["rows"][0]["ip_address"] == ""
    ids = extract_asset_identifiers(hostname_parsed["rows"][0], scanner_source="openvas")
    assert any(item["identifier_type"] == "HOSTNAME" and item["normalized_value"] == "db01" for item in ids)


def test_openvas_solution_type_only_vendor_fix_sets_patch_available():
    expected = {
        "VendorFix": "1",
        "Vendor Fix": "1",
        "Workaround": "0",
        "Mitigation": "0",
        "NoneAvailable": "0",
        "WillNotFix": "0",
    }
    for solution_type, patch_available in expected.items():
        payload = f"""<?xml version='1.0'?><get_reports_response><report><report><results>
        <result id='solution-type'><host>192.0.2.44<hostname>greenbone.example.test</hostname></host>
        <severity>9.8</severity><nvt><name>Greenbone solution semantics</name>
        <refs><ref type='cve' id='CVE-2026-65010'/></refs>
        <solution type='{solution_type}'>Remediation guidance is present.</solution></nvt></result>
        </results></report></report></get_reports_response>""".encode()
        parsed = parse_import_file(payload, filename=f"{solution_type}.xml")
        assert parsed["rows"][0]["patch_available"] == patch_available


def test_openvas_xml_delta_history_is_not_imported_as_current_finding():
    content = b"""<?xml version='1.0' encoding='UTF-8'?>
    <report><results start='1' max='1'>
      <result id='current-result'>
        <name>Current finding</name>
        <host>192.0.2.10<hostname>host.example.test</hostname></host>
        <port>443/tcp</port>
        <nvt oid='1.3.6.1.4.1.25623.1.0.100001'>
          <name>Current NVT</name><cvss_base>9.8</cvss_base>
          <refs><ref type='cve' id='CVE-2026-10001'/></refs>
        </nvt>
        <severity>9.8</severity><description>Current vulnerable result</description>
        <delta>changed
          <result id='previous-result'>
            <name>Previous finding</name>
            <host>192.0.2.10<hostname>host.example.test</hostname></host>
            <port>443/tcp</port>
            <nvt oid='1.3.6.1.4.1.25623.1.0.100002'>
              <name>Previous NVT</name><cvss_base>7.5</cvss_base>
              <refs><ref type='cve' id='CVE-2025-9999'/></refs>
            </nvt>
            <severity>7.5</severity><description>Historical delta comparison result</description>
          </result>
          <diff>changed</diff>
        </delta>
      </result>
    </results></report>
    """
    parsed = parse_import_file(content, filename="greenbone-delta.xml")
    assert parsed["detected_format"] == "openvas_xml"
    assert parsed["metadata"]["result_count"] == 1
    assert [row["cve_id"] for row in parsed["rows"]] == ["CVE-2026-10001"]
    assert all(row["cve_id"] != "CVE-2025-9999" for row in parsed["rows"])


def test_openvas_xml_preserves_greenbone_asset_uuid_across_host_changes(tmp_path: Path):
    asset_uuid = "3f90bda4-1ca6-4f08-b9a7-e88b2a2e52c8"

    def payload(result_id: str, ip: str, hostname: str) -> bytes:
        return f"""<?xml version='1.0'?><report><results><result id='{result_id}'>
        <name>Stable Greenbone asset</name>
        <host>{ip}<asset asset_id='{asset_uuid}'/><hostname>{hostname}</hostname></host>
        <port>443/tcp</port><nvt oid='1.3.6.1.4.1.25623.1.0.77777'>
        <name>Stable Greenbone asset</name><cvss_base>8.0</cvss_base>
        <refs><ref type='cve' id='CVE-2026-77777'/></refs></nvt>
        <severity>8.0</severity><description>Identity continuity regression</description>
        </result></results></report>""".encode()

    first_parsed = parse_import_file(payload("r1", "192.0.2.10", "old.example.test"), filename="first.xml")
    second_parsed = parse_import_file(payload("r2", "192.0.2.11", "new.example.test"), filename="second.xml")
    assert first_parsed["rows"][0]["asset_id"] == asset_uuid
    assert second_parsed["rows"][0]["asset_id"] == asset_uuid

    identifiers = extract_asset_identifiers(first_parsed["rows"][0], scanner_source="openvas")
    assert any(
        item["identifier_type"] == "SCANNER_ASSET_ID" and item["normalized_value"] == asset_uuid
        for item in identifiers
    )

    db = tmp_path / "greenbone-asset-identity.sqlite3"
    init_db(db)
    previous_db_path = main.DB_PATH
    main.DB_PATH = db
    try:
        first_row = dict(first_parsed["rows"][0])
        first_row["finding_id"] = "GB-ASSET-R1"
        second_row = dict(second_parsed["rows"][0])
        second_row["finding_id"] = "GB-ASSET-R2"
        first = main.normalize_row(first_row, 0, scanner_source="openvas")
        second = main.normalize_row(second_row, 1, scanner_source="openvas")
        apply_import_batch(db, [first], scanner_source="openvas", filename="first.xml")
        result = apply_import_batch(db, [second], scanner_source="openvas", filename="second.xml")
    finally:
        main.DB_PATH = previous_db_path

    assert result["inserted"] == 0
    assert result["updated"] == 1
    assert result["merged"] == 1
    assert len(list_findings(db)) == 1
    assert len(list_assets(db)) == 1


def test_openvas_csv_solution_type_semantics_and_legacy_fallback():
    typed = (
        "IP,Hostname,NVT Name,CVEs,CVSS,Solution,Solution Type\n"
        "192.0.2.50,typed.example.test,Typed,CVE-2026-65011,8.0,Temporary workaround,Workaround\n"
        "192.0.2.51,vendor.example.test,Vendor,CVE-2026-65012,8.0,Install vendor patch,VendorFix\n"
        "192.0.2.52,none.example.test,None,CVE-2026-65013,8.0,No fix currently exists,NoneAvailable\n"
    ).encode()
    parsed = parse_import_file(typed, filename="typed-greenbone.csv")
    assert [row["patch_available"] for row in parsed["rows"]] == ["0", "1", "0"]

    legacy = (
        "Host;Hostname;Port;NVT Name;CVEs;Severity;Summary;Solution\n"
        "198.51.100.30;;22/tcp;Legacy SSH issue;CVE-2026-65014;7.2;Legacy export;Upgrade SSH\n"
    ).encode("utf-8-sig")
    legacy_parsed = parse_import_file(legacy, filename="legacy-greenbone.csv", format_hint="openvas")
    assert legacy_parsed["rows"][0]["patch_available"] == "1"

def test_preview_session_is_actor_bound(tmp_path: Path):
    token = create_preview_session(
        tmp_path,
        content=b"product,cve_id\nDemo,CVE-2026-55555\n",
        filename="demo.csv",
        format_hint="auto",
        actor="operator-a",
        ttl_seconds=1800,
    )
    metadata, content = load_preview_session(tmp_path, token, actor="operator-a", ttl_seconds=1800)
    assert metadata["filename"] == "demo.csv"
    assert b"CVE-2026-55555" in content
    try:
        load_preview_session(tmp_path, token, actor="operator-b", ttl_seconds=1800)
    except PermissionError:
        pass
    else:
        raise AssertionError("preview session must be actor-bound")


def test_import_preview_and_apply_openvas_csv(client: TestClient, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(main, "IMPORT_PREVIEW_DIR", tmp_path / "previews")
    token = _csrf(client)
    content = (
        "IP,Hostname,NVT Name,CVEs,CVSS,Summary\n"
        "10.0.0.21,app-21,Demo NVT,CVE-2026-61001,9.4,Patch now\n"
    ).encode()
    preview = client.post(
        "/upload/findings/preview",
        data={"csrf_token": token, "format_hint": "auto", "scanner_source": "", "import_mode": "incremental"},
        files={"file": ("openvas.csv", content, "text/csv")},
    )
    assert preview.status_code == 200
    assert "OpenVAS/Greenbone CSV" in preview.text
    assert "반영 가능" in preview.text
    match = re.search(r'name="token" value="([A-Za-z0-9_-]+)"', preview.text)
    assert match
    preview_token = match.group(1)
    apply = client.post(
        "/upload/findings/apply",
        data={
            "csrf_token": token,
            "token": preview_token,
            "scanner_source": "openvas-lab",
            "import_mode": "incremental",
        },
        follow_redirects=False,
    )
    assert apply.status_code == 303
    found = client.get("/api/v1/findings?query=CVE-2026-61001").json()["items"]
    assert len(found) == 1
    assert found[0]["scanner_source"] == "openvas-lab"
    assert found[0]["asset_name"] == "app-21"


def test_invalid_rows_can_be_downloaded_and_explicitly_skipped(client: TestClient, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(main, "IMPORT_PREVIEW_DIR", tmp_path / "previews")
    token = _csrf(client)
    content = (
        "product,cve_id,asset_name,cvss\n"
        "Good,CVE-2026-62001,host-good,8.0\n"
        "Bad,NOT-A-CVE,host-bad,7.0\n"
    ).encode()
    preview = client.post(
        "/upload/findings/preview",
        data={"csrf_token": token, "format_hint": "csv", "scanner_source": "manual", "import_mode": "incremental"},
        files={"file": ("mixed.csv", content, "text/csv")},
    )
    assert preview.status_code == 200
    assert "확인 필요한 항목" in preview.text
    match = re.search(r'name="token" value="([A-Za-z0-9_-]+)"', preview.text)
    assert match
    preview_token = match.group(1)
    common = {
        "csrf_token": token,
        "token": preview_token,
        "scanner_source": "manual",
        "import_mode": "incremental",
    }
    errors = client.post("/upload/findings/errors", data=common)
    assert errors.status_code == 200
    assert errors.content.startswith(b"\xef\xbb\xbf")
    assert b"NOT-A-CVE" in errors.content
    blocked = client.post("/upload/findings/apply", data=common)
    assert blocked.status_code == 400
    applied = client.post(
        "/upload/findings/apply",
        data={**common, "skip_invalid": "1"},
        follow_redirects=False,
    )
    assert applied.status_code == 303
    assert "upload_partial" in applied.headers["location"]
    assert len(client.get("/api/v1/findings?query=CVE-2026-62001").json()["items"]) == 1



def test_invalid_ip_is_reported_during_preview_before_apply(client: TestClient, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(main, "IMPORT_PREVIEW_DIR", tmp_path / "previews")
    token = _csrf(client)
    content = (
        "product,cve_id,asset_name,ip_address\n"
        "Good,CVE-2026-77700,db00,192.0.2.10\n"
        "Bad,CVE-2026-77701,db01,db01\n"
    ).encode()
    preview = client.post(
        "/upload/findings/preview",
        data={"csrf_token": token, "format_hint": "csv", "scanner_source": "manual", "import_mode": "incremental"},
        files={"file": ("bad-ip.csv", content, "text/csv")},
    )
    assert preview.status_code == 200
    assert "확인 필요한 항목" in preview.text
    assert "IP 주소 형식이 올바르지 않습니다: db01" in preview.text
    match = re.search(r'name="token" value="([A-Za-z0-9_-]+)"', preview.text)
    assert match
    blocked = client.post(
        "/upload/findings/apply",
        data={
            "csrf_token": token,
            "token": match.group(1),
            "scanner_source": "manual",
            "import_mode": "incremental",
        },
    )
    assert blocked.status_code == 400


def test_generic_csv_bracketed_ipv6_previews_and_applies(client: TestClient, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(main, "IMPORT_PREVIEW_DIR", tmp_path / "previews")
    token = _csrf(client)
    content = (
        "product,cve_id,asset_name,ip_address\n"
        "Demo,CVE-2026-77702,[2001:db8::1],[2001:db8::1]\n"
    ).encode()
    preview = client.post(
        "/upload/findings/preview",
        data={"csrf_token": token, "format_hint": "csv", "scanner_source": "manual", "import_mode": "incremental"},
        files={"file": ("ipv6.csv", content, "text/csv")},
    )
    assert preview.status_code == 200
    assert "반영 가능" in preview.text
    match = re.search(r'name="token" value="([A-Za-z0-9_-]+)"', preview.text)
    assert match
    applied = client.post(
        "/upload/findings/apply",
        data={
            "csrf_token": token,
            "token": match.group(1),
            "scanner_source": "manual",
            "import_mode": "incremental",
        },
        follow_redirects=False,
    )
    assert applied.status_code == 303
    found = client.get("/api/v1/findings?query=CVE-2026-77702").json()["items"]
    assert len(found) == 1


def test_openvas_bracketed_ipv6_does_not_create_false_hostname_identity():
    content = (
        "IP,NVT Name,CVEs,CVSS\n"
        "[2001:db8::1],IPv6 target,CVE-2026-77703,7.5\n"
    ).encode()
    parsed = parse_import_file(content, filename="openvas-ipv6.csv", format_hint="openvas")
    identifiers = extract_asset_identifiers(parsed["rows"][0], scanner_source="openvas")
    assert any(
        item["identifier_type"] == "IP_ADDRESS" and item["normalized_value"] == "2001:db8::1"
        for item in identifiers
    )
    assert not any(
        item["identifier_type"] == "HOSTNAME" and item["normalized_value"] == "[2001:db8::1]"
        for item in identifiers
    )

def test_snapshot_import_never_skips_invalid_rows(client: TestClient, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(main, "IMPORT_PREVIEW_DIR", tmp_path / "previews")
    token = _csrf(client)
    content = b"product,cve_id\nGood,CVE-2026-63001\nBad,INVALID\n"
    preview = client.post(
        "/upload/findings/preview",
        data={"csrf_token": token, "format_hint": "csv", "scanner_source": "snapshot-test", "import_mode": "snapshot"},
        files={"file": ("snapshot.csv", content, "text/csv")},
    )
    match = re.search(r'name="token" value="([A-Za-z0-9_-]+)"', preview.text)
    assert match
    response = client.post(
        "/upload/findings/apply",
        data={
            "csrf_token": token,
            "token": match.group(1),
            "scanner_source": "snapshot-test",
            "import_mode": "snapshot",
            "skip_invalid": "1",
        },
    )
    assert response.status_code == 400
    assert "전체 결과 대조에서는 오류 행 건너뛰기를 허용하지 않습니다" in response.text



def test_nessus_cpe22_cvss4_and_asset_uuid_are_preserved():
    payload = (Path(__file__).parent / "fixtures" / "scanners" / "nessus-cpe22.nessus").read_bytes()
    parsed = parse_import_file(payload, filename="nessus-cpe22.nessus")
    row = parsed["rows"][0]
    assert row["product"] == "acme widget_server"
    assert row["product_version"] == "1.2.3"
    assert row["cvss"] == "9.3"
    assert row["asset_id"] == "fixture-host-uuid"
    assert parsed["metadata"]["cpe22_items"] == 1
    assert parsed["metadata"]["adapter_profile"] == "nessus-client-data-v2"



def test_nessus_smbios_absent_uuid_sentinels_never_false_merge_assets(tmp_path: Path):
    def payload(ip: str, fqdn: str, bios_uuid: str, cve: str, *, mcafee_guid: str = "") -> bytes:
        mcafee = f"<tag name='mcafee-epo-guid'>{mcafee_guid}</tag>" if mcafee_guid else ""
        return f"""<?xml version='1.0'?><NessusClientData_v2><Report name='demo'>
        <ReportHost name='{ip}'><HostProperties><tag name='host-ip'>{ip}</tag>
        <tag name='host-fqdn'>{fqdn}</tag><tag name='bios-uuid'>{bios_uuid}</tag>{mcafee}</HostProperties>
        <ReportItem port='443' protocol='tcp' pluginID='991000' pluginName='SMBIOS UUID identity'>
        <cve>{cve}</cve><cvss3_base_score>7.5</cvss3_base_score></ReportItem>
        </ReportHost></Report></NessusClientData_v2>""".encode()

    zero = "00000000-0000-0000-0000-000000000000"
    ff = "FFFFFFFF-FFFF-FFFF-FFFF-FFFFFFFFFFFF"
    first_parsed = parse_import_file(
        payload("192.0.2.10", "alpha.example.test", zero, "CVE-2026-88881"),
        filename="zero-bios-a.nessus",
    )
    second_parsed = parse_import_file(
        payload("198.51.100.20", "beta.example.test", zero, "CVE-2026-88882"),
        filename="zero-bios-b.nessus",
    )
    assert first_parsed["rows"][0]["asset_id"] == ""
    assert second_parsed["rows"][0]["asset_id"] == ""
    assert any("bios-uuid" in warning and "sentinel" in warning for warning in first_parsed["parser_warnings"])

    ff_parsed = parse_import_file(
        payload("203.0.113.30", "ff.example.test", ff, "CVE-2026-88883"),
        filename="ff-bios.nessus",
    )
    assert ff_parsed["rows"][0]["asset_id"] == ""

    fallback_guid = "EPO-GUID-1234"
    fallback_parsed = parse_import_file(
        payload("203.0.113.31", "epo.example.test", zero, "CVE-2026-88884", mcafee_guid=fallback_guid),
        filename="zero-bios-epo.nessus",
    )
    assert fallback_parsed["rows"][0]["asset_id"] == fallback_guid

    valid_bios = "9F5C2CB8-4FA8-4A0B-9E66-3C7C0B9D0A77"
    valid_parsed = parse_import_file(
        payload("203.0.113.32", "valid.example.test", valid_bios, "CVE-2026-88885"),
        filename="valid-bios.nessus",
    )
    assert valid_parsed["rows"][0]["asset_id"] == valid_bios

    db = tmp_path / "nessus-bios-sentinel.sqlite3"
    init_db(db)
    previous_db_path = main.DB_PATH
    main.DB_PATH = db
    try:
        for index, parsed in enumerate((first_parsed, second_parsed)):
            row = dict(parsed["rows"][0])
            row["finding_id"] = f"NESSUS-BIOS-{index + 1}"
            normalized = main.normalize_row(row, index, scanner_source="nessus")
            apply_import_batch(db, [normalized], scanner_source="nessus", filename=f"bios-{index}.nessus")
    finally:
        main.DB_PATH = previous_db_path

    assert len(list_assets(db)) == 2
    assert len(list_findings(db)) == 2


def test_nessus_has_patch_false_overrides_solution_text():
    payload = b"""<?xml version='1.0'?><NessusClientData_v2><Report name='demo'>
    <ReportHost name='host1.example.test'><HostProperties><tag name='host-ip'>192.0.2.10</tag></HostProperties>
    <ReportItem port='443' protocol='tcp' pluginID='990001' pluginName='No vendor patch'>
    <cve>CVE-2026-65020</cve><cvss3_base_score>7.5</cvss3_base_score>
    <has_patch>false</has_patch><solution>Use compensating controls until a vendor fix exists.</solution>
    </ReportItem></ReportHost></Report></NessusClientData_v2>"""
    parsed = parse_import_file(payload, filename="has-patch-false.nessus")
    assert parsed["rows"][0]["patch_available"] == "0"


def test_nessus_has_patch_true_is_authoritative_even_without_solution_text():
    payload = b"""<?xml version='1.0'?><NessusClientData_v2><Report name='demo'>
    <ReportHost name='host2.example.test'><HostProperties><tag name='host-ip'>192.0.2.11</tag></HostProperties>
    <ReportItem port='0' protocol='tcp' pluginID='990002' pluginName='Vendor patch exists'>
    <cve>CVE-2026-65021</cve><cvss3_base_score>8.1</cvss3_base_score><has_patch>true</has_patch>
    </ReportItem></ReportHost></Report></NessusClientData_v2>"""
    parsed = parse_import_file(payload, filename="has-patch-true.nessus")
    assert parsed["rows"][0]["patch_available"] == "1"


def test_nessus_has_patch_invalid_is_fail_closed_and_legacy_export_keeps_fallback():
    invalid = b"""<?xml version='1.0'?><NessusClientData_v2><Report name='demo'>
    <ReportHost name='host3.example.test'><ReportItem port='0' protocol='tcp' pluginID='990003' pluginName='Malformed patch flag'>
    <cve>CVE-2026-65022</cve><has_patch>maybe</has_patch><solution>Upgrade the affected package.</solution>
    </ReportItem></ReportHost></Report></NessusClientData_v2>"""
    parsed = parse_import_file(invalid, filename="has-patch-invalid.nessus")
    assert parsed["rows"][0]["patch_available"] == "0"
    assert any("has_patch boolean" in warning for warning in parsed["parser_warnings"])

    legacy = b"""<?xml version='1.0'?><NessusClientData_v2><Report name='demo'>
    <ReportHost name='host4.example.test'><ReportItem port='0' protocol='tcp' pluginID='990004' pluginName='Legacy patch guidance'>
    <cve>CVE-2026-65023</cve><solution>Upgrade the affected package.</solution>
    </ReportItem></ReportHost></Report></NessusClientData_v2>"""
    legacy_parsed = parse_import_file(legacy, filename="legacy-no-has-patch.nessus")
    assert legacy_parsed["rows"][0]["patch_available"] == "1"

def test_openvas_xml_extracts_cve_ref_attributes_and_solution():
    payload = (Path(__file__).parent / "fixtures" / "scanners" / "openvas-refs.xml").read_bytes()
    parsed = parse_import_file(payload, filename="openvas-refs.xml")
    row = parsed["rows"][0]
    assert row["cve_id"] == "CVE-2026-96002"
    assert row["patch_available"] == "1"
    assert "해결 방법" in row["notes"]
    assert parsed["metadata"]["attribute_ref_results"] == 1


def test_extensionless_utf8_bom_openvas_xml_is_detected():
    payload = (Path(__file__).parent / "fixtures" / "scanners" / "openvas-refs.xml").read_bytes()
    parsed = parse_import_file(b"\xef\xbb\xbf" + payload, filename="customer-export")
    assert parsed["detected_format"] == "openvas_xml"
    assert parsed["rows"][0]["cve_id"] == "CVE-2026-96002"


def test_openvas_semicolon_csv_uses_host_as_ip_address():
    payload = (Path(__file__).parent / "fixtures" / "scanners" / "openvas-semicolon.csv").read_bytes()
    parsed = parse_import_file(payload, filename="openvas-semicolon.csv")
    row = parsed["rows"][0]
    assert parsed["detected_format"] == "openvas_csv"
    assert parsed["metadata"]["delimiter"] == ";"
    assert row["asset_name"] == "198.51.100.30"
    assert row["ip_address"] == "198.51.100.30"
    assert row["patch_available"] == "1"


@pytest.mark.parametrize(
    ("payload", "message"),
    [
        (
            b"<?xml version='1.0'?><!DOCTYPE report [<!ENTITY x 'boom'>]><report>&x;</report>",
            "DOCTYPE 또는 ENTITY",
        ),
        (
            ("<get_reports_response>" + "<x>" * 129 + "</x>" * 129 + "</get_reports_response>").encode(),
            "중첩 깊이",
        ),
    ],
)
def test_scanner_xml_structural_guards_block_unsafe_documents(payload: bytes, message: str):
    with pytest.raises(ValueError, match=message):
        parse_import_file(payload, filename="unsafe.xml")


def test_duplicate_scanner_preview_surfaces_parser_warning(client: TestClient, tmp_path: Path, monkeypatch):
    monkeypatch.setattr(main, "IMPORT_PREVIEW_DIR", tmp_path / "previews")
    token = _csrf(client)
    payload = (Path(__file__).parent / "fixtures" / "scanners" / "openvas-duplicate.xml").read_bytes()
    preview = client.post(
        "/upload/findings/preview",
        data={
            "csrf_token": token,
            "format_hint": "auto",
            "scanner_source": "openvas-duplicate",
            "import_mode": "incremental",
        },
        files={"file": ("openvas-duplicate.xml", payload, "application/xml")},
    )
    assert preview.status_code == 200
    assert "스캐너 파일 호환성: REVIEW" in preview.text
    assert "파서 경고" in preview.text
    assert "중복" in preview.text


def test_duplicate_header_suffix_collisions_never_overwrite_csv_or_xlsx_columns():
    csv_content = (
        "product,cve_id,notes,notes,notes_2\n"
        "Demo,CVE-2026-77801,A,B,C\n"
    ).encode()
    parsed_csv = parse_import_file(csv_content, filename="header-collision.csv", format_hint="csv")
    assert parsed_csv["headers"] == ["product", "cve_id", "notes", "notes_2", "notes_2_2"]
    assert parsed_csv["rows"][0] == {
        "product": "Demo", "cve_id": "CVE-2026-77801", "notes": "A", "notes_2": "B", "notes_2_2": "C"
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["product", "cve_id", "notes", "notes", "notes_2"])
    sheet.append(["Demo", "CVE-2026-77802", "A", "B", "C"])
    payload = io.BytesIO()
    workbook.save(payload)
    parsed_xlsx = parse_import_file(payload.getvalue(), filename="header-collision.xlsx", format_hint="xlsx")
    assert parsed_xlsx["headers"] == ["product", "cve_id", "notes", "notes_2", "notes_2_2"]
    assert parsed_xlsx["rows"][0]["notes_2"] == "B"
    assert parsed_xlsx["rows"][0]["notes_2_2"] == "C"


def test_invalid_explicit_fqdn_is_rejected_by_asset_identity_normalizer():
    for value in (
        "bad host.example.com",
        "bad/host.example.com",
        "-leading.example.com",
        "trailing-.example.com",
        f"{'a' * 64}.example.com",
        "192.0.2.10",
    ):
        with pytest.raises(ValueError, match="FQDN 형식이 올바르지 않습니다"):
            normalize_asset_identifier("FQDN", value)
    assert normalize_asset_identifier("FQDN", "DB01.Example.COM.") == "db01.example.com"
    assert normalize_asset_identifier("FQDN", "dead.beef") == "dead.beef"


def test_scanner_source_case_change_reuses_same_source_record(tmp_path: Path):
    db = tmp_path / "scanner-source-case.sqlite3"
    init_db(db)
    row = {
        "finding_id": "CASE-1",
        "product": "Case scanner",
        "product_version": "1.0",
        "asset_id": "CASE-ASSET-1",
        "asset_name": "case.example.test",
        "environment": "prod",
        "cve_id": "CVE-2026-77803",
        "component": "case-component",
        "component_version": "1.0",
        "cvss": 7.5,
        "epss": 0.1,
        "epss_percentile": 0.2,
        "kev": 0,
        "internet_exposed": 0,
        "asset_criticality": 3,
        "data_sensitivity": 3,
        "patch_available": 1,
        "compensating_control": 0,
        "status": "OPEN",
        "owner": "",
        "due_date": "",
        "notes": "",
        "score": 50,
        "threat_score": 20,
        "asset_context_score": 20,
        "remediation_urgency_score": 10,
        "decision": "REVIEW",
        "decision_label": "검토",
        "sla_days": 30,
        "target_date": "2026-09-01",
        "mitigation_required": 0,
        "reasons": "test",
        "policy_version": "test",
        "first_seen_at": "2026-08-12",
        "first_scored_at": "2026-08-12",
        "last_scored_at": "2026-08-12",
        "record_state": "ACTIVE",
        "row_version": 1,
    }
    first = apply_import_batch(db, [row], scanner_source="Nessus-DMZ", filename="one.csv")
    second = apply_import_batch(db, [row], scanner_source="nessus-dmz", filename="two.csv")
    assert first["inserted"] == 1
    assert second["updated"] == 1
    detail = get_source_reconciliation(db, "CASE-1")
    assert len(detail["records"]) == 1
    assert detail["records"][0]["scanner_source"] == "nessus-dmz"


def test_openvas_same_nvt_cve_on_multiple_ports_imports_as_distinct_findings(tmp_path: Path):
    xml_payload = b"""<?xml version='1.0'?><report><results>
    <result id='r443'><host>192.0.2.88<asset asset_id='GB-MULTIPORT-ASSET'/><hostname>app.example.test</hostname></host>
    <port>443/tcp</port><nvt oid='1.3.6.1.4.1.25623.1.0.88888'><name>TLS component vulnerability</name><refs><ref type='cve' id='CVE-2026-88888'/></refs></nvt></result>
    <result id='r8443'><host>192.0.2.88<asset asset_id='GB-MULTIPORT-ASSET'/><hostname>app.example.test</hostname></host>
    <port>8443/tcp</port><nvt oid='1.3.6.1.4.1.25623.1.0.88888'><name>TLS component vulnerability</name><refs><ref type='cve' id='CVE-2026-88888'/></refs></nvt></result>
    </results></report>"""
    csv_payload = (
        "IP,Hostname,Port,NVT Name,CVEs,CVSS,Summary\n"
        "192.0.2.89,csv.example.test,443/tcp,TLS component vulnerability,CVE-2026-88889,8.8,one\n"
        "192.0.2.89,csv.example.test,8443/tcp,TLS component vulnerability,CVE-2026-88889,8.8,two\n"
    ).encode()

    for filename, payload in (("multi-port.xml", xml_payload), ("multi-port.csv", csv_payload)):
        parsed = parse_import_file(payload, filename=filename)
        assert [row["component"] for row in parsed["rows"]] == [
            "TLS component vulnerability [443/tcp]",
            "TLS component vulnerability [8443/tcp]",
        ]
        mapped, _, errors = map_import_rows(parsed["rows"], parsed["source_rows"], parsed["mapping"])
        assert errors == []
        rows = []
        for index, row in enumerate(mapped):
            prepared = dict(row)
            prepared["finding_id"] = f"OPENVAS-PORT-{filename}-{index + 1}"
            rows.append(main.normalize_row(prepared, index, scanner_source="openvas"))
        db = tmp_path / f"{filename}.sqlite3"
        init_db(db)
        result = apply_import_batch(db, rows, scanner_source="openvas", filename=filename)
        assert result["inserted"] == 2
        assert len(list_findings(db)) == 2

    host_level_xml = b"""<?xml version='1.0'?><report><results><result id='host'>
    <host>192.0.2.90</host><port>0/tcp</port><nvt oid='1.3.6.1.4.1.25623.1.0.88890'><name>Host-level vulnerability</name><refs><ref type='cve' id='CVE-2026-88890'/></refs></nvt>
    </result></results></report>"""
    assert parse_import_file(host_level_xml, filename="host-level.xml")["rows"][0]["component"] == "Host-level vulnerability"
    general_xml = host_level_xml.replace(b"0/tcp", b"general/tcp")
    assert parse_import_file(general_xml, filename="general.xml")["rows"][0]["component"] == "Host-level vulnerability"



def test_openvas_modern_port_protocol_csv_header_preserves_multi_port_identity(tmp_path: Path):
    payload = (
        "IP,Hostname,Port/Protocol,NVT Name,CVEs,CVSS,Summary\n"
        "192.0.2.111,modern.example.test,443/tcp,TLS component vulnerability,CVE-2026-90111,8.8,one\n"
        "192.0.2.111,modern.example.test,8443/tcp,TLS component vulnerability,CVE-2026-90111,8.8,two\n"
    ).encode()
    parsed = parse_import_file(payload, filename="modern-greenbone.csv", format_hint="openvas")
    assert [row["component"] for row in parsed["rows"]] == [
        "TLS component vulnerability [443/tcp]",
        "TLS component vulnerability [8443/tcp]",
    ]
    mapped, _, errors = map_import_rows(parsed["rows"], parsed["source_rows"], parsed["mapping"])
    assert errors == []
    rows = []
    for index, row in enumerate(mapped):
        prepared = dict(row)
        prepared["finding_id"] = f"OPENVAS-MODERN-PORT-{index + 1}"
        rows.append(main.normalize_row(prepared, index, scanner_source="openvas"))
    db = tmp_path / "openvas-modern-port-protocol.sqlite3"
    init_db(db)
    result = apply_import_batch(db, rows, scanner_source="openvas", filename="modern-greenbone.csv")
    assert result["inserted"] == 2
    assert len(list_findings(db)) == 2

def test_openvas_customizable_csv_split_port_protocol_and_vt_name_imports_distinct_findings(tmp_path: Path):
    payload = (
        "IP,Hostname,Port,Port Protocol,VT Name,CVEs,Severity,Severity Level,Summary\n"
        "192.0.2.120,custom.example.test,443,tcp,TLS component vulnerability,CVE-2026-90120,8.8,High,one\n"
        "192.0.2.120,custom.example.test,443,udp,TLS component vulnerability,CVE-2026-90120,8.8,High,two\n"
    ).encode()
    parsed = parse_import_file(payload, filename="customizable-greenbone.csv")
    assert parsed["detected_format"] == "openvas_csv"
    assert [row["product"] for row in parsed["rows"]] == [
        "TLS component vulnerability",
        "TLS component vulnerability",
    ]
    assert [row["component"] for row in parsed["rows"]] == [
        "TLS component vulnerability [443/tcp]",
        "TLS component vulnerability [443/udp]",
    ]
    mapped, _, errors = map_import_rows(parsed["rows"], parsed["source_rows"], parsed["mapping"])
    assert errors == []
    rows = []
    for index, row in enumerate(mapped):
        prepared = dict(row)
        prepared["finding_id"] = f"OPENVAS-CUSTOM-PORT-PROTO-{index + 1}"
        rows.append(main.normalize_row(prepared, index, scanner_source="openvas"))
    db = tmp_path / "openvas-customizable-port-protocol.sqlite3"
    init_db(db)
    result = apply_import_batch(db, rows, scanner_source="openvas", filename="customizable-greenbone.csv")
    assert result["inserted"] == 2
    assert len(list_findings(db)) == 2


def test_greenbone_current_security_intelligence_csv_cve_references_imports_distinct_findings(tmp_path: Path):
    payload = (
        "Vulnerability name,Port/Protocol,CVE references,Host name,IP address,Severity,Solution type,Solution,QoD,Summary,Impact\n"
        "TLS component vulnerability,443/tcp,CVE-2026-90200,current.example.test,192.0.2.200,8.8,VendorFix,Upgrade,95,one,impact one\n"
        "TLS component vulnerability,8443/tcp,CVE-2026-90200,current.example.test,192.0.2.200,8.8,VendorFix,Upgrade,95,two,impact two\n"
    ).encode()
    parsed = parse_import_file(payload, filename="vulnerabilities_with_affected_assets.csv")
    assert parsed["detected_format"] == "openvas_csv"
    assert parsed["adapter"] == "openvas"
    assert [row["cve_id"] for row in parsed["rows"]] == ["CVE-2026-90200", "CVE-2026-90200"]
    assert [row["product"] for row in parsed["rows"]] == [
        "TLS component vulnerability",
        "TLS component vulnerability",
    ]
    assert [row["component"] for row in parsed["rows"]] == [
        "TLS component vulnerability [443/tcp]",
        "TLS component vulnerability [8443/tcp]",
    ]
    mapped, _, errors = map_import_rows(parsed["rows"], parsed["source_rows"], parsed["mapping"])
    assert errors == []
    rows = []
    for index, row in enumerate(mapped):
        prepared = dict(row)
        prepared["finding_id"] = f"GREENBONE-CURRENT-CSV-{index + 1}"
        rows.append(main.normalize_row(prepared, index, scanner_source="openvas"))
    db = tmp_path / "greenbone-current-security-intelligence.sqlite3"
    init_db(db)
    result = apply_import_batch(
        db,
        rows,
        scanner_source="openvas",
        filename="vulnerabilities_with_affected_assets.csv",
    )
    assert result["inserted"] == 2
    assert len(list_findings(db)) == 2


def test_greenbone_current_detailed_csv_preserves_epss_and_percentile(tmp_path: Path):
    payload = (
        "Severity,EPSS score,EPSS percentile,Vulnerability name,Solution type,Solution,QoD,Summary,Impact,Specific result,CVE references,Port/Protocol,Host name,IP address\n"
        "8.8,0.82,0.99,TLS component vulnerability,Vendor fix,Upgrade,95,one,impact,detail,CVE-2026-90300,443/tcp,epss.example.test,192.0.2.210\n"
    ).encode()
    parsed = parse_import_file(payload, filename="vulnerabilities_with_affected_assets.csv")
    assert parsed["detected_format"] == "openvas_csv"
    assert parsed["adapter"] == "openvas"
    assert parsed["rows"][0]["epss"] == "0.82"
    assert parsed["rows"][0]["epss_percentile"] == "0.99"

    mapped, _, errors = map_import_rows(parsed["rows"], parsed["source_rows"], parsed["mapping"])
    assert errors == []
    prepared = dict(mapped[0])
    prepared["finding_id"] = "GREENBONE-CURRENT-EPSS-1"
    normalized = main.normalize_row(prepared, 0, scanner_source="openvas")
    assert normalized["epss"] == 0.82
    assert normalized["epss_percentile"] == 0.99

    db = tmp_path / "greenbone-current-epss.sqlite3"
    init_db(db)
    result = apply_import_batch(
        db,
        [normalized],
        scanner_source="openvas",
        filename="vulnerabilities_with_affected_assets.csv",
    )
    assert result["inserted"] == 1
    finding = list_findings(db)[0]
    assert finding["epss"] == 0.82
    assert finding["epss_percentile"] == 0.99


def test_nessus_same_plugin_cve_on_multiple_ports_imports_as_distinct_findings(tmp_path: Path):
    payload = b"""<?xml version='1.0'?><NessusClientData_v2><Report name='multi-port'>
    <ReportHost name='app.example.test'><HostProperties>
    <tag name='host-ip'>192.0.2.55</tag><tag name='host-fqdn'>app.example.test</tag>
    <tag name='host-uuid'>HOST-PORT-UUID</tag></HostProperties>
    <ReportItem port='443' protocol='tcp' svc_name='https' pluginID='55555' pluginName='TLS component vulnerability'>
    <cve>CVE-2026-55555</cve><cvss3_base_score>8.8</cvss3_base_score></ReportItem>
    <ReportItem port='8443' protocol='tcp' svc_name='https-alt' pluginID='55555' pluginName='TLS component vulnerability'>
    <cve>CVE-2026-55555</cve><cvss3_base_score>8.8</cvss3_base_score></ReportItem>
    </ReportHost></Report></NessusClientData_v2>"""
    parsed = parse_import_file(payload, filename="multi-port.nessus")
    assert [row["component"] for row in parsed["rows"]] == [
        "TLS component vulnerability [443/tcp]",
        "TLS component vulnerability [8443/tcp]",
    ]

    host_level = b"""<?xml version='1.0'?><NessusClientData_v2><Report name='host-level'>
    <ReportHost name='app.example.test'><HostProperties><tag name='host-uuid'>HOST-PORT-UUID</tag></HostProperties>
    <ReportItem port='0' protocol='tcp' svc_name='general' pluginID='55556' pluginName='Host-level vulnerability'>
    <cve>CVE-2026-55556</cve></ReportItem></ReportHost></Report></NessusClientData_v2>"""
    assert parse_import_file(host_level, filename="host-level.nessus")["rows"][0]["component"] == "Host-level vulnerability"

    mapped, _, errors = map_import_rows(parsed["rows"], parsed["source_rows"], parsed["mapping"])
    assert errors == []
    rows = []
    for index, row in enumerate(mapped):
        prepared = dict(row)
        prepared["finding_id"] = f"NESSUS-PORT-{index + 1}"
        rows.append(main.normalize_row(prepared, index, scanner_source="nessus"))

    db = tmp_path / "nessus-multi-port.sqlite3"
    init_db(db)
    result = apply_import_batch(db, rows, scanner_source="nessus", filename="multi-port.nessus")
    findings = list_findings(db)
    assert result["inserted"] == 2
    assert len(findings) == 2
    assert {item["component"] for item in findings} == {
        "TLS component vulnerability [443/tcp]",
        "TLS component vulnerability [8443/tcp]",
    }
