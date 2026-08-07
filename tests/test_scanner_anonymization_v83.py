from __future__ import annotations

import io
import json
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook

import app.main as main
from app.services.finding_imports import parse_import_file
from app.services.scanner_anonymization import anonymize_scanner_file, build_scanner_collection_bundle

FIXTURES = Path(__file__).parent / "fixtures" / "scanners"


def _bundle_entries(payload: bytes) -> dict[str, bytes]:
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def test_nessus_bundle_removes_original_identifiers_and_stays_parseable():
    payload = b"""<?xml version='1.0'?><NessusClientData_v2><Report name='corp-scan' customer='Acme Secret'>
    <ReportHost name='db01.corp.internal'><HostProperties>
    <tag name='host-ip'>10.23.45.67</tag><tag name='host-fqdn'>db01.corp.internal</tag>
    <tag name='host-uuid'>123e4567-e89b-42d3-a456-426614174000</tag>
    </HostProperties><ReportItem port='443' protocol='tcp' svc_name='https' pluginID='9001' pluginName='TLS issue'>
    <cve>CVE-2026-99101</cve><cvss3_base_score>8.8</cvss3_base_score>
    <plugin_output>Contact secops@corp.internal on db01.corp.internal.</plugin_output>
    <solution>Upgrade the affected package.</solution><custom_customer>Acme Custom Metadata</custom_customer></ReportItem></ReportHost></Report></NessusClientData_v2>"""
    bundle, summary = build_scanner_collection_bundle(payload, filename="secret-customer.nessus")
    entries = _bundle_entries(bundle)
    assert summary["compatibility_status"] in {"READY", "REVIEW"}
    assert set(entries) == {
        "README.txt",
        "reports/anonymization.json",
        "reports/compatibility.json",
        "sample/sanitized-scanner-sample.nessus",
    }
    combined = b"\n".join(entries.values())
    for secret in (b"secret-customer", b"corp-scan", b"Acme Secret", b"Acme Custom Metadata", b"db01.corp.internal", b"10.23.45.67", b"secops@corp.internal", b"123e4567-e89b-42d3-a456-426614174000"):
        assert secret not in combined
    parsed = parse_import_file(entries["sample/sanitized-scanner-sample.nessus"], filename="sample.nessus")
    assert parsed["rows"][0]["cve_id"] == "CVE-2026-99101"
    assert parsed["rows"][0]["ip_address"].startswith(("192.0.2.", "198.51.100.", "203.0.113."))
    report = json.loads(entries["reports/anonymization.json"])
    assert report["mapping_included"] is False
    assert report["residual_source_identifiers"] == []


def test_openvas_xml_repeated_host_uses_consistent_alias():
    payload = b"""<get_reports_response><report><report><results>
    <owner><name>Acme Customer</name></owner><custom>Hidden Tenant</custom>
    <result><name>One</name><host>172.16.10.5<hostname>app.corp.local</hostname></host><port>443/tcp</port>
    <nvt><name>One</name><cve>CVE-2026-99102</cve><cvss_base>7.5</cvss_base></nvt><description>secret one</description></result>
    <result><name>Two</name><host>172.16.10.5<hostname>app.corp.local</hostname></host><port>80/tcp</port>
    <nvt><name>Two</name><cve>CVE-2026-99103</cve><cvss_base>6.5</cvss_base></nvt><description>secret two</description></result>
    </results></report></report></get_reports_response>"""
    result = anonymize_scanner_file(payload, filename="customer.xml")
    root = ET.fromstring(result["content"])
    hosts = [str(item.text or "").strip() for item in root.iter() if item.tag.rsplit("}", 1)[-1].casefold() == "host"]
    hostnames = [str(item.text or "").strip() for item in root.iter() if item.tag.rsplit("}", 1)[-1].casefold() == "hostname"]
    assert len(set(hosts)) == 1
    assert len(set(hostnames)) == 1
    assert b"172.16.10.5" not in result["content"]
    assert b"app.corp.local" not in result["content"]
    assert b"Acme Customer" not in result["content"]
    assert b"Hidden Tenant" not in result["content"]
    assert result["compatibility"]["importable_rows"] == 2


def test_csv_anonymization_preserves_headers_and_cves():
    content = (
        "IP,Hostname,NVT Name,CVEs,CVSS,Summary,Owner,Internal Code\n"
        "10.2.3.4,web.corp.internal,TLS issue,CVE-2026-99104,9.1,Internal path /srv/prod,alice@corp.internal,ACME-SECRET-42\n"
    ).encode()
    result = anonymize_scanner_file(content, filename="greenbone.csv")
    text = result["content"].decode("utf-8-sig")
    assert text.splitlines()[0] == "IP,Hostname,NVT Name,CVEs,CVSS,Summary,Owner,Internal Code"
    assert "CVE-2026-99104" in text
    assert "10.2.3.4" not in text
    assert "web.corp.internal" not in text
    assert "alice@corp.internal" not in text
    assert "/srv/prod" not in text
    assert "ACME-SECRET-42" not in text
    assert result["compatibility"]["importable_rows"] == 1


def test_xlsx_anonymization_rebuilds_values_only_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customer PROD"
    sheet.append(["product", "cve_id", "asset_name", "ip_address", "notes"])
    sheet.append(["OpenSSL", "CVE-2026-99105", "api.corp.internal", "10.20.30.40", "owner bob@corp.internal"])
    sheet["A2"].hyperlink = "https://internal.corp.local/package"
    source = io.BytesIO()
    workbook.save(source)
    result = anonymize_scanner_file(source.getvalue(), filename="secret.xlsx")
    target = load_workbook(io.BytesIO(result["content"]), read_only=False, data_only=True)
    try:
        assert target.sheetnames == ["Sheet1"]
        row = [target["Sheet1"].cell(2, column).value for column in range(1, 6)]
        assert row[1] == "CVE-2026-99105"
        assert row[2] != "api.corp.internal"
        assert row[3] != "10.20.30.40"
        assert row[4] == "Redacted by VulnFlow anonymizer."
        assert target["Sheet1"]["A2"].hyperlink is None
    finally:
        target.close()


def test_strict_profile_pseudonymizes_product_and_cpe():
    payload = (FIXTURES / "nessus-cpe22.nessus").read_bytes()
    result = anonymize_scanner_file(payload, filename="customer.nessus", profile="strict")
    text = result["content"].decode("utf-8")
    assert "acme" not in text.casefold()
    assert "widget_server" not in text.casefold()
    assert re.search(r"sample-(?:product|component)-\d{4}", text)
    assert result["compatibility"]["importable_rows"] == 1


def test_cli_bundle_contains_no_source_mapping(tmp_path: Path):
    source = FIXTURES / "openvas-refs.xml"
    output = tmp_path / "bundle.zip"
    from scripts.scanner_collection_bundle import main
    import sys

    old = sys.argv
    try:
        sys.argv = ["scanner_collection_bundle.py", str(source), "--output", str(output)]
        assert main() == 0
    finally:
        sys.argv = old
    entries = _bundle_entries(output.read_bytes())
    assert not any("mapping" in name.casefold() for name in entries)
    assert source.name.encode() not in b"\n".join(entries.values())


def test_web_endpoint_returns_nonpersistent_collection_bundle(client):
    page = client.get("/upload")
    assert page.status_code == 200
    token = client.cookies.get(main.CSRF_COOKIE)
    payload = (FIXTURES / "openvas-refs.xml").read_bytes()
    response = client.post(
        "/upload/findings/anonymize",
        data={"csrf_token": token, "format_hint": "auto", "profile": "compatibility"},
        files={"file": ("customer-secret.xml", payload, "application/xml")},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/zip")
    assert "customer-secret" not in response.headers["content-disposition"]
    entries = _bundle_entries(response.content)
    assert "reports/anonymization.json" in entries
    assert b"customer-secret" not in b"\n".join(entries.values())
